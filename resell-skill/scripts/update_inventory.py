#!/usr/bin/env python3
"""Manage the resell inventory tracker spreadsheet.

Usage:
    python3 update_inventory.py <xlsx_path> add --name "Item" --category "Cat" --price-low 50 --price-high 100 --price-recommended 75
    python3 update_inventory.py <xlsx_path> update --name "Item" --status sold --sold-price 80
    python3 update_inventory.py <xlsx_path> list [--status listed]
    python3 update_inventory.py <xlsx_path> init

Creates/updates an Excel spreadsheet tracking all items in the resale pipeline.
"""

import sys
import argparse
from pathlib import Path
from datetime import datetime

def init_workbook(xlsx_path):
    """Create a new inventory workbook with proper formatting."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    headers = [
        'Item Name', 'Category', 'Brand', 'Model', 'Condition',
        'Quick Sale $', 'Market Price $', 'Optimistic $', 'Listed Price $',
        'Sold Price $', 'Marketplace', 'Status', 'Date Added', 'Date Listed',
        'Date Sold', 'Listing URL', 'Photos Folder', 'Notes'
    ]

    # Header styling
    header_fill = PatternFill(start_color='1B4F72', end_color='1B4F72', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    thin_border = Border(
        bottom=Side(style='thin', color='CCCCCC')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Column widths
    widths = [25, 20, 15, 20, 12, 12, 12, 12, 12, 12, 18, 12, 12, 12, 12, 35, 30, 30]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col) if col <= 26 else 'A' + chr(64 + col - 26)].width = width

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Add a summary sheet
    summary = wb.create_sheet("Summary")
    summary['A1'] = 'Resell Inventory Summary'
    summary['A1'].font = Font(size=14, bold=True)
    summary['A3'] = 'Total Items:'
    summary['A4'] = 'Listed:'
    summary['A5'] = 'Sold:'
    summary['A6'] = 'Draft:'
    summary['A7'] = 'Total Revenue:'
    summary['A8'] = 'Avg Margin vs Quick Sale:'

    # Formulas will reference Inventory sheet
    summary['B3'] = '=COUNTA(Inventory!A2:A1000)'
    summary['B4'] = '=COUNTIF(Inventory!L2:L1000,"listed")'
    summary['B5'] = '=COUNTIF(Inventory!L2:L1000,"sold")'
    summary['B6'] = '=COUNTIF(Inventory!L2:L1000,"draft")'
    summary['B7'] = '=SUM(Inventory!J2:J1000)'

    wb.save(xlsx_path)
    print(f"Initialized inventory at {xlsx_path}")
    return wb


def add_item(xlsx_path, **kwargs):
    """Add an item to the inventory."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = load_workbook(xlsx_path)
    ws = wb['Inventory']

    next_row = ws.max_row + 1

    # Map kwargs to columns
    col_map = {
        'name': 1, 'category': 2, 'brand': 3, 'model': 4, 'condition': 5,
        'price_low': 6, 'price_mid': 7, 'price_high': 8, 'listed_price': 9,
        'sold_price': 10, 'marketplace': 11, 'status': 12,
        'date_added': 13, 'date_listed': 14, 'date_sold': 15,
        'listing_url': 16, 'photos_folder': 17, 'notes': 18
    }

    # Set defaults
    if 'date_added' not in kwargs:
        kwargs['date_added'] = datetime.now().strftime('%Y-%m-%d')
    if 'status' not in kwargs:
        kwargs['status'] = 'draft'

    for key, col in col_map.items():
        if key in kwargs and kwargs[key] is not None:
            ws.cell(row=next_row, column=col, value=kwargs[key])

    # Status color coding
    status = kwargs.get('status', 'draft')
    status_colors = {
        'draft': 'FFF3CD',
        'listed': 'D4EDDA',
        'sold': 'D1ECF1',
        'expired': 'F8D7DA'
    }
    if status in status_colors:
        ws.cell(row=next_row, column=12).fill = PatternFill(
            start_color=status_colors[status],
            end_color=status_colors[status],
            fill_type='solid'
        )

    wb.save(xlsx_path)
    print(f"Added '{kwargs.get('name', 'item')}' to inventory (row {next_row})")


def update_item(xlsx_path, name, **kwargs):
    """Update an existing item by name."""
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    wb = load_workbook(xlsx_path)
    ws = wb['Inventory']

    col_map = {
        'status': 12, 'listed_price': 9, 'sold_price': 10,
        'marketplace': 11, 'date_listed': 14, 'date_sold': 15,
        'listing_url': 16, 'notes': 18
    }

    # Find the row
    found = False
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == name:
            for key, col in col_map.items():
                if key in kwargs and kwargs[key] is not None:
                    ws.cell(row=row, column=col, value=kwargs[key])

            # Update status color
            status = kwargs.get('status')
            if status:
                status_colors = {
                    'draft': 'FFF3CD', 'listed': 'D4EDDA',
                    'sold': 'D1ECF1', 'expired': 'F8D7DA'
                }
                if status in status_colors:
                    ws.cell(row=row, column=12).fill = PatternFill(
                        start_color=status_colors[status],
                        end_color=status_colors[status],
                        fill_type='solid'
                    )

            found = True
            print(f"Updated '{name}' in inventory")
            break

    if not found:
        print(f"Item '{name}' not found in inventory")
        return

    wb.save(xlsx_path)


def list_items(xlsx_path, status_filter=None):
    """List items in the inventory."""
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    ws = wb['Inventory']

    items = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if name is None:
            continue
        status = ws.cell(row=row, column=12).value
        if status_filter and status != status_filter:
            continue
        items.append({
            'name': name,
            'category': ws.cell(row=row, column=2).value,
            'market_price': ws.cell(row=row, column=7).value,
            'status': status,
            'marketplace': ws.cell(row=row, column=11).value,
        })

    if not items:
        print("No items found")
    else:
        for item in items:
            price_str = f"${item['market_price']}" if item['market_price'] else "unpriced"
            print(f"  {item['name']} | {item['category']} | {price_str} | {item['status']} | {item['marketplace'] or '-'}")

    return items


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Manage resell inventory')
    parser.add_argument('xlsx_path', help='Path to inventory spreadsheet')
    subparsers = parser.add_subparsers(dest='command')

    # init
    subparsers.add_parser('init')

    # add
    add_p = subparsers.add_parser('add')
    add_p.add_argument('--name', required=True)
    add_p.add_argument('--category')
    add_p.add_argument('--brand')
    add_p.add_argument('--model')
    add_p.add_argument('--condition')
    add_p.add_argument('--price-low', type=float, dest='price_low')
    add_p.add_argument('--price-mid', type=float, dest='price_mid')
    add_p.add_argument('--price-high', type=float, dest='price_high')
    add_p.add_argument('--marketplace')
    add_p.add_argument('--status', default='draft')
    add_p.add_argument('--photos-folder', dest='photos_folder')
    add_p.add_argument('--notes')

    # update
    upd_p = subparsers.add_parser('update')
    upd_p.add_argument('--name', required=True)
    upd_p.add_argument('--status')
    upd_p.add_argument('--listed-price', type=float, dest='listed_price')
    upd_p.add_argument('--sold-price', type=float, dest='sold_price')
    upd_p.add_argument('--marketplace')
    upd_p.add_argument('--listing-url', dest='listing_url')
    upd_p.add_argument('--notes')

    # list
    list_p = subparsers.add_parser('list')
    list_p.add_argument('--status')

    args = parser.parse_args()

    if args.command == 'init':
        init_workbook(args.xlsx_path)
    elif args.command == 'add':
        add_item(args.xlsx_path, **{k: v for k, v in vars(args).items() if k not in ('xlsx_path', 'command')})
    elif args.command == 'update':
        update_item(args.xlsx_path, args.name, **{k: v for k, v in vars(args).items() if k not in ('xlsx_path', 'command', 'name')})
    elif args.command == 'list':
        list_items(args.xlsx_path, args.status)
